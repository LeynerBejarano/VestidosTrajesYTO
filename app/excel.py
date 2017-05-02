import os
from app import app
from app.model.pedido import Pedido
from datetime import datetime, date
from pytz import timezone
import openpyxl as xlsx
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from flask import jsonify, send_from_directory, request
from flask.ext.login import login_required
from app.model.cargo import Cargo
from app.model.ciudad import Ciudad
from app.model.evento import Evento
from app.model.jornada import Jornada
from app.model.nivele import Nivele
from app.model.prenda import Prenda
from app.model.color import Color
from app.model.estilo import Estilo
from app.model.accesorio import Accesorio
from app.model.cliente import Cliente
from app.model.empresa import Empresa
from app.model.user import Usuario
from app.model.institucion import Institucion
from app.model.pedido import Pedido
from app.model.det_pedido import Det_pedido
from app.model.prospecto import Prospecto
from app.model.det_prospecto import Det_prospecto
from app.model.clase import Clase
from app.model.presindiv import Presindiv
from app.model.prespedido import Prespedido
from app.model.prestamo import Prestamo
from app.model.despacho import Despacho
from app.model.det_despacho import Det_despacho
from app.model.det_estola import Det_estola
from app.model.pro_estola import Pro_estola
from app.model.tipo import Tipo
from app.model.talla import Talla
from app.model.entrega import Entrega
from app.model.tipo_estola import Tipo_estola
from app.model.terminacion import Terminacion
from app.model.presentacion import Presentacion
from app.model.estado_com import Estado_com
from app.model.estado_fin import Estado_fin
from app.model.estado import Estado
from app.model.orden import Orden
from app.model.personalizada import Personalizada
from app.model.transportadora import Transportadora
from app.model.tipo_orden import Tipo_orden
from app.model.abono import Abono
from app.model.tipo_pago import Tipo_pago

def crear_excel(filename, temporada = None, activo = False):
    wb = xlsx.Workbook()
    # grab the active worksheet
    ws = wb.active

    excel = open(os.path.join(app.config['UPLOAD_FOLDER'], 'excel/' + filename), 'wb')
    cabecera_pedido = ['No. pedido', 'No. pedido manual', 'fecha del pedido', 'Mes', 'niños /adultos', 'tipo de evento', '# de togas', '# de cortesías', 'color de togas', 'color birrete', 'estilo togas', 'estilo birrete', 'Valor Unitario $', 'Valor total del pedido $', '# de birretes','# de borlas', 'color  borlas', 'Tipo de estolas', '# de estolas', 'color estolas', 'accesorios', '# de accesorios', 'color accesorios', 'petos', '# de petos', 'fecha del evento día/mes', 'Fecha de entrega', 'Hora de entrega', 'Vendedor', 'Estado']
    cabecera_cliente = ['nombre del encargado', 'Cargo', 'C.C.', 'Cumpleaños día/mes', 'Celular del encargado', 'tel. Fijo del encargado', 'correo electrónico del encargado', 'Dirección del encargado',   'barrio del encargado', 'Municipio del encargado'] 
    cabecera_pedido += cabecera_cliente
    cabecera_institucion = ['Institución/Nombre Empresa', 'nit', 'municipio de la institución',  'am/fam', 'nivel educativo', 'jornada']
    cabecera_pedido += cabecera_institucion
    ws.append(cabecera_pedido)

    pedidos = Pedido.query.all()
    for pedido in pedidos:
        fecha_evento = pedido.ped_fecha_evento
        if (not temporada and not activo) or (not activo and temporada == 'A' and fecha_evento and fecha_evento >= date(fecha_evento.year, 11,1)) or (not activo and temporada == 'B' and fecha_evento and fecha_evento < date(fecha_evento.year, 11,1)) or (not temporada and activo and pedido.ped_estado_com == 1) or (activo and pedido.ped_estado_com == 1 and temporada == 'A' and fecha_evento and fecha_evento >= date(fecha_evento.year, 11,1)) or (activo and pedido.ped_estado_com == 1 and temporada == 'B' and fecha_evento and fecha_evento < date(fecha_evento.year, 11,1)):
            if pedido.ped_poblacion == 1:
                poblacion = 'adultos'
            else:
                poblacion = 'niños'
            detalles = {}
            for detalle in Det_pedido.query.filter(Det_pedido.det_pedido == pedido.ped_numero).all():
                detalles[detalle.det_clase] = detalle
            
            datos_pedido = [pedido.ped_numero, 
                            pedido.ped_manual, 
                            pedido.ped_fecha_mod.strftime('%m/%d/%Y'), 
                            pedido.ped_fecha_mod.strftime('%m'), 
                            poblacion, 
                            Evento.query.get(pedido.ped_evento).eve_nombre, 
                            check_detalle(1, detalles, 'detalles[1].det_pedida', 0), 
                            check_detalle(1, detalles, 'detalles[1].det_cortesia', 0), 
                            check_detalle(1, detalles, 'Color.query.get(detalles[1].det_color).col_nombre', ''), 
                            check_detalle(2, detalles, 'Color.query.get(detalles[2].det_color).col_nombre', ''), 
                            check_detalle(1, detalles, 'Estilo.query.get(detalles[1].det_estilo).est_nombre', ''), 
                            check_detalle(2, detalles, 'Estilo.query.get(detalles[2].det_estilo).est_nombre', ''), 
                            pedido.ped_val_unitario, 
                            check_detalle(1, detalles, 'detalles[1].det_pedida', 0) * pedido.ped_val_unitario, 
                            check_detalle(2, detalles, 'detalles[2].det_pedida', 0), 
                            check_detalle(3, detalles, 'detalles[3].det_pedida', 0), 
                            check_detalle(3, detalles, 'Color.query.get(detalles[3].det_color).col_nombre', ''), 
                            check_detalle(4, detalles, 'tipo_estola(detalles[4])', 'No'), 
                            check_detalle(4, detalles, 'detalles[4].det_pedida', 0), 
                            check_detalle(4, detalles, 'Color.query.get(detalles[4].det_color).col_nombre', ''), 
                            get_accesorio(detalles), 
                            get_accesorio_cant(detalles), 
                            get_accesorio_col(detalles), 
                            check_value(8 in detalles.keys(), "'Sí'", 'No'), 
                            check_detalle(8, detalles, 'detalles[8].det_pedida', 0), 
                            check_value(pedido.ped_fecha_evento, "pedido.ped_fecha_evento.strftime('%d/%m/%Y')", '', data=['pedido', pedido]),
                            check_value(pedido.ped_fecha_entrega, "pedido.ped_fecha_entrega.strftime('%d/%m/%Y')", '', data=['pedido', pedido]), 
                            check_value(pedido.ped_hora_entrega, "pedido.ped_hora_entrega.strftime('%I:%M %p')", '', data=['pedido', pedido]),
                            Usuario.query.get(pedido.ped_vendedor).usu_nombre + ' ' + Usuario.query.get(pedido.ped_vendedor).usu_apellido, 
                            Estado_com.query.get(pedido.ped_estado_com).esc_nombre ]

            cliente = Cliente.query.get(pedido.ped_cliente)
            datos_cliente =[cliente.cli_nombre + ' ' + cliente.cli_apellido, 
                            Cargo.query.get(cliente.cli_cargo).crg_descripcion,
                            cliente.cli_identificacion,
                            check_value(cliente.cli_nacido_dia and cliente.cli_nacido_mes, "str(cliente.cli_nacido_dia) + '-' + to_month(cliente.cli_nacido_mes)", '', data=['cliente', cliente]),
                            cliente.cli_celular,
                            cliente.cli_telefono,
                            cliente.cli_email,
                            cliente.cli_direccion,
                            cliente.cli_barrio,
                            Ciudad.query.get(cliente.cli_ciudad).ciu_nombre]
            datos_pedido += datos_cliente

            institucion = Institucion.query.get(pedido.ped_institucion)
            datos_institucion =[institucion.ins_nombre,
                                str(institucion.ins_nit),
                                Ciudad.query.get(institucion.ins_ciudad).ciu_nombre,
                                area_metropol(institucion.ins_ciudad),
                                Nivele.query.get(pedido.ped_nivel).niv_nombre,
                                Jornada.query.get(pedido.ped_jornada).jor_nombre]
            datos_pedido += datos_institucion
            ws.append(datos_pedido)

    rows = ws.max_row - 1
    fill = PatternFill("solid", fgColor="dbdbdb")
    style_range(ws, 'A1:AD' + str(rows + 1), fill=fill)

    fill = PatternFill("solid", fgColor="d9e1f2")
    style_range(ws, 'AE:AN' + str(rows + 1), fill=fill)

    fill = PatternFill("solid", fgColor="b4c6e7")
    style_range(ws, 'AO:AT' + str(rows + 1), fill=fill)
    # Save the file
    wb.save(excel)
    excel.close()


def crear_excel_abonos(filename):
    wb = xlsx.Workbook()
    # grab the active worksheet
    ws = wb.active

    excel = open(os.path.join(app.config['UPLOAD_FOLDER'], 'excel/' + filename), 'wb')
    cabecera_abono = ['Abono No.', 'Pedido No.', 'Valor', 'Tipo de pago', 'Fecha de pago (d/m/a)', 'Fecha de creación (d/m/a)']
    ws.append(cabecera_abono)

    abonos = Abono.query.all()
    for abono in abonos:            
        datos_abono =  [abono.abo_id,
                        abono.abo_pedido,
                        abono.abo_valor,
                        Tipo_pago.query.get(abono.abo_tipo).tip_nombre,
                        check_value(abono.abo_fecha, "abono.abo_fecha.strftime('%d/%m/%Y')", '', data=['abono', abono]),
                        abono.abo_fecha_mod.strftime('%d/%m/%Y')]

        ws.append(datos_abono)

    # Save the file
    wb.save(excel)
    excel.close()

def crear_excel_estolas(filename):
    wb = xlsx.Workbook()
    # grab the active worksheet
    ws = wb.active

    excel = open(os.path.join(app.config['UPLOAD_FOLDER'], 'excel/' + filename), 'wb')
    cabecera_estola = ['No. pedido', 'Institución', 'Encargado', 'Celular', 'tel fijo', 'valor unitario', '# de estolas', 'valor total del pedido', 'color estola', 'acabado', 'Terminación', 'Presentación', 'Tamaño', 'Doble faz', 'Flequillo', 'Personalizada','fecha de entrega', 'hora de entrega', 'fecha del evento']
    ws.append(cabecera_estola)

    estolas = Det_pedido.query.filter(Det_pedido.det_clase == 4).all()

    tam_estolas = {1: 'Ancha', 2: 'Normal', 3: 'Estrecha'}
    total_estolas_color = {}
    total_estolas = 0
    for estola in estolas:
        pedido = Pedido.query.get(estola.det_pedido)
        institucion = Institucion.query.get(pedido.ped_institucion)
        cliente = Cliente.query.get(pedido.ped_cliente)
        info_estola = Det_estola.query.filter(Det_estola.etl_detalle == estola.det_id).first()
        fecha_evento = pedido.ped_fecha_evento
        if (pedido.ped_estado_com == 1 and fecha_evento and fecha_evento >= date(fecha_evento.year, 11,1) and info_estola.etl_tipo == 2):

            #### Se trae la toga con la cual se calcula el valor total del pedido
            #### Esto hay que cambiarlo cuando se habilite la prenda principal en el pedido
            detalle = Det_pedido.query.filter(Det_pedido.det_pedido == pedido.ped_numero, Det_pedido.det_clase == pedido.ped_principal).first()

            total_estolas += estola.det_pedida
            color = Color.query.get(estola.det_color).col_nombre
            if color in total_estolas_color.keys():
                total_estolas_color[color] += estola.det_pedida
            else:
                total_estolas_color[color] = estola.det_pedida

            datos_estola = [pedido.ped_numero, 
                            institucion.ins_nombre, 
                            cliente.cli_nombre, 
                            cliente.cli_celular, 
                            cliente.cli_telefono, 
                            pedido.ped_val_unitario, 
                            estola.det_pedida,
                            check_value(detalle, "detalle.det_pedida", 0, data=['detalle', detalle]) * pedido.ped_val_unitario,
                            color, 
                            Estilo.query.get(estola.det_estilo).est_nombre, 
                            Terminacion.query.get(info_estola.etl_terminacion).ter_nombre, 
                            Presentacion.query.get(info_estola.etl_presentacion).prs_nombre, 
                            tam_estolas[info_estola.etl_tamano], 
                            check_value(info_estola.etl_doble_faz == 1, "'Si'", 'No'), 
                            check_value(info_estola.etl_flequillo == 1, "'Si'", 'No'), 
                            check_value(info_estola.etl_personalizada == 1, "'Si'", 'No'),  
                            check_value(pedido.ped_fecha_entrega, "pedido.ped_fecha_entrega.strftime('%d/%m/%Y')", '', data=['pedido', pedido]),
                            check_value(pedido.ped_hora_entrega, "pedido.ped_hora_entrega.strftime('%H:%M %p')", '', data=['pedido', pedido]),
                            check_value(pedido.ped_fecha_evento, "pedido.ped_fecha_evento.strftime('%d/%m/%Y')", '', data=['pedido', pedido])]

            ws.append(datos_estola)

    ws.append([])
    ws.append(['# TOTAL DE ESTOLAS', '', total_estolas])
    first = True
    for key, value in total_estolas_color.items():
        if first: 
            ws.append(['# Estolas por color', key, value])
            first = False
        else:
            ws.append(['', key, value])

    # Save the file
    wb.save(excel)
    excel.close()

def crear_excel_orden_entrega_recogida(filename, fecha):
    wb = xlsx.Workbook()
    # grab the active worksheet
    ws = wb.active

    excel = open(os.path.join(app.config['UPLOAD_FOLDER'], 'excel/' + filename), 'wb')

    cabecera_entrega = ['No. pedido', 'Institución', 'Encargado', 'Celular', 'tel fijo','Cantidad','color toga', 'tipo estola', 'color estola', 'Motivo', 'tipo de entrega', 'hora de entrega', 'Total abonado']
    ws.append(cabecera_entrega)

    entregas = Orden.query.filter(Orden.ord_fecha == fecha, Orden.ord_tipo == 1 ).order_by(Orden.ord_hora).all()
    for entrega in entregas:
        pedido = Pedido.query.get(entrega.ord_pedido)
        despacho = Despacho.query.get(entrega.ord_despacho)
        institucion = Institucion.query.get(pedido.ped_institucion)
        cliente = Cliente.query.get(pedido.ped_cliente)
        fecha_evento = pedido.ped_fecha_evento
        toga = Det_pedido.query.filter(Det_pedido.det_pedido == pedido.ped_numero, Det_pedido.det_clase == pedido.ped_principal).first()
        estola = Det_pedido.query.filter(Det_pedido.det_pedido == pedido.ped_numero, Det_pedido.det_clase == 4).first()
        if estola:
            det_estola = Det_estola.query.filter(Det_estola.etl_detalle == estola.det_id).first()
        cantidad = 0
        for detalle in Det_despacho.query.filter(Det_despacho.tal_despacho == despacho.des_id):
            cantidad += detalle.tal_cantidad
        datos_entrega = [pedido.ped_numero, 
                        institucion.ins_nombre, 
                        cliente.cli_nombre + " " + cliente.cli_apellido, 
                        cliente.cli_celular, 
                        cliente.cli_telefono, 
                        cantidad,
                        Color.query.get(toga.det_color).col_nombre if toga else '',
                        Tipo.query.get(det_estola.etl_tipo).tip_nombre if det_estola else '',
                        Color.query.get(estola.det_color).col_nombre if estola else '',
                        Prestamo.query.get(despacho.des_prestamo).prs_nombre,
                        Tipo_orden.query.get(entrega.ord_tipo_orden).tip_nombre, 
                        entrega.ord_hora.strftime('%I:%M %p') if entrega.ord_hora else '',
                        pedido.ped_abono]

        ws.append(datos_entrega)

    # Save the file
    wb.save(excel)
    excel.close()


@app.route('/exportar_excel')
@login_required
def exportar_excel():
    filename = 'pedidos_' + datetime.now(timezone('America/Bogota')).strftime('%Y-%m-%d') + '.xlsx'
    crear_excel(filename)
    return send_from_directory(app.config['UPLOAD_FOLDER'] + 'excel/', filename, as_attachment=True)

@app.route('/exportar_excel_A')
@login_required
def exportar_excel_a():
    filename = 'pedidos_' + datetime.now(timezone('America/Bogota')).strftime('%Y-%m-%d') + '_A.xlsx'
    crear_excel(filename, temporada = 'A')
    return send_from_directory(app.config['UPLOAD_FOLDER'] + 'excel/', filename, as_attachment=True)

@app.route('/exportar_excel_B')
@login_required
def exportar_excel_b():
    filename = 'pedidos_' + datetime.now(timezone('America/Bogota')).strftime('%Y-%m-%d') + '_B.xlsx'
    crear_excel(filename, temporada = 'B')
    return send_from_directory(app.config['UPLOAD_FOLDER'] + 'excel/', filename, as_attachment=True)

@app.route('/exportar_excel_activos')
@login_required
def exportar_excel_activos():
    filename = 'pedidos-activos_' + datetime.now(timezone('America/Bogota')).strftime('%Y-%m-%d') + '.xlsx'
    crear_excel(filename, activo = True)
    return send_from_directory(app.config['UPLOAD_FOLDER'] + 'excel/', filename, as_attachment=True)

@app.route('/exportar_excel_activos_A')
@login_required
def exportar_excel_activos_A():
    filename = 'pedidos-activos_' + datetime.now(timezone('America/Bogota')).strftime('%Y-%m-%d') + '_A.xlsx'
    crear_excel(filename, activo = True, temporada = 'A')
    return send_from_directory(app.config['UPLOAD_FOLDER'] + 'excel/', filename, as_attachment=True)

@app.route('/exportar_excel_activos_B')
@login_required
def exportar_excel_activos_B():
    filename = 'pedidos-activos_' + datetime.now(timezone('America/Bogota')).strftime('%Y-%m-%d') + '_B.xlsx'
    crear_excel(filename, activo = True, temporada = 'B')
    return send_from_directory(app.config['UPLOAD_FOLDER'] + 'excel/', filename, as_attachment=True)

@app.route('/exportar_excel_abonos')
@login_required
def exportar_excel_abonos():
    filename = 'abonos_' + datetime.now(timezone('America/Bogota')).strftime('%Y-%m-%d') + '.xlsx'
    crear_excel_abonos(filename)
    return send_from_directory(app.config['UPLOAD_FOLDER'] + 'excel/', filename, as_attachment=True)

@app.route('/exportar_excel_estolas')
@login_required
def exportar_excel_estolas():
    filename = 'estolas_' + datetime.now(timezone('America/Bogota')).strftime('%Y-%m-%d') + '.xlsx'
    crear_excel_estolas(filename)
    return send_from_directory(app.config['UPLOAD_FOLDER'] + 'excel/', filename, as_attachment=True)

@app.route('/exportar_excel_entrega_recogida')
@login_required
def exportar_excel_entrega_recogida():
    fecha = request.args.get('fecha')
    filename = 'orden_entrega_recogida_' + fecha + '.xlsx'
    fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
    crear_excel_orden_entrega_recogida(filename, fecha)
    return send_from_directory(app.config['UPLOAD_FOLDER'] + 'excel/', filename, as_attachment=True)



def check_detalle(llave, detalles, verdadero, falso):
    if llave in detalles.keys():
        return eval(verdadero)
    else:
        return falso

def check_value(valor, verdadero, falso, data=None):
    if data:
        exec(data[0] + ' = data[1]')
    if valor:
        return eval(verdadero)
    else:
        return falso

def tipo_estola(detalle):
    if detalle:
        estola = Det_estola.query.filter(Det_estola.etl_detalle == detalle.det_id).first()
        if estola.etl_tipo == 2:
            return Tipo_estola.query.get(estola.etl_tipo_escudo).tes_nombre
        else:
            return 'En alquiler'
    else:
        return 'No'

def get_accesorio(detalles):
    accesorios = ''
    for i in range(5, 8):
        if i in detalles.keys():
            return Clase.query.get(i).cla_nombre
    return 'Ninguno'

def get_accesorio_cant(detalles):
    accesorios = ''
    for i in range(5, 8):
        if i in detalles.keys():
            return detalles[i].det_pedida
    return ''

def get_accesorio_col(detalles):
    accesorios = ''
    for i in range(5, 8):
        if i in detalles.keys():
            return Color.query.get(detalles[i].det_color).col_nombre
    return ''

def to_month(mes):
    meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    return meses[mes - 1]

def area_metropol(ciudad):
    if Ciudad.query.get(ciudad).ciu_metropol == 1:
        return 'AM'
    else:
        return 'FAM'



def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill